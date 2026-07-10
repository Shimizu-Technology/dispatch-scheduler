#!/usr/bin/env python3
"""Create public-safe demo seed data from local workflow artifacts.

The source workbooks stay outside Git. This importer preserves useful operating
shape (counts, priorities, statuses, trades, regions, dates, and crew sizes)
while replacing people, locations, work-order identifiers, descriptions, and
source references before writing the committed demo seed.
"""
from __future__ import annotations

import argparse
import json
import os
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_ARTIFACT_DIR = ROOT / "docs" / "examples-from-john"
ARTIFACT_DIR = Path(os.environ.get("JOHN_ARTIFACT_DIR", DEFAULT_ARTIFACT_DIR))
OUT = ROOT / "data" / "seeds" / "sample_data.json"

PUBLIC_SANITIZATION_VERSION = 2
IMPORT_CONFIG = {"driver_names": [], "location_regions": {}}

PUBLIC_DESCRIPTIONS = {
    "Plumbing": "Inspect and repair a plumbing issue at the listed demo site.",
    "HVAC": "Inspect and repair an HVAC or refrigeration issue at the listed demo site.",
    "Electrical": "Inspect and repair an electrical issue at the listed demo site.",
    "Painting": "Assess and complete painting work at the listed demo site.",
    "Carpentry": "Assess and complete carpentry or building work at the listed demo site.",
    "Landscaping": "Complete landscaping work at the listed demo site.",
    "Masonry": "Assess and complete masonry work at the listed demo site.",
    "General": "Assess and complete general facilities work at the listed demo site.",
}

APPROVED_WORK_ORDER_COLUMNS = {
    "client": 0,
    "location": 1,
    "wo_number": 2,
    "description": 3,
    "team_name": 4,
    "approval_status": 6,
    "dispatch_status": 7,
}

SKILL_ALIASES = {
    "HVAC": "HVAC",
    "ELECTRICAL": "Electrical",
    "PLUMBING": "Plumbing",
    "CARPENTRY": "Carpentry",
    "CARPENTER": "Carpentry",
    "MASON": "Masonry",
    "MASONRY": "Masonry",
    "LANDSCAPING": "Landscaping",
    "PAINTING": "Painting",
    "GENERAL": "General",
    "HELPER": "Helper",
}

STATUS_FIXES = {
    "APPRVOED": "APPROVED",
    "ASSESSMNET": "ASSESSMENT",
    "ASSESSMET": "ASSESSMENT",
    "ASSESMENT": "ASSESSMENT",
}


def clean(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).replace("\n", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text or None


def canonical(text: str | None) -> str:
    value = (text or "").upper().strip()
    for typo, fixed in STATUS_FIXES.items():
        value = value.replace(typo, fixed)
    return value


def find_required_file(patterns: list[str], env_var: str | None = None) -> Path:
    if env_var and os.environ.get(env_var):
        selected = Path(os.environ[env_var])
        return selected if selected.is_absolute() else ARTIFACT_DIR / selected

    matches = []
    for pattern in patterns:
        matches.extend(ARTIFACT_DIR.glob(pattern))
    matches = sorted(set(matches))
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        names = ", ".join(path.name for path in matches)
        hint = f" Set {env_var} to choose one." if env_var else ""
        raise ValueError(f"Multiple files match {', '.join(patterns)} in {ARTIFACT_DIR}: {names}.{hint}")

    expected = ", ".join(patterns)
    raise FileNotFoundError(f"Could not find {expected} in {ARTIFACT_DIR}")


def find_required_workbook(required_sheets: list[str], env_var: str, excluded_sheets: list[str] | None = None) -> Path:
    if os.environ.get(env_var):
        return find_required_file(["*.xlsx"], env_var=env_var)

    required = {name.strip().lower() for name in required_sheets}
    excluded = {name.strip().lower() for name in (excluded_sheets or [])}
    matches = []
    for path in sorted(ARTIFACT_DIR.glob("*.xlsx")):
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_names = {name.strip().lower() for name in workbook.sheetnames}
        workbook.close()
        if required.issubset(sheet_names) and sheet_names.isdisjoint(excluded):
            matches.append(path)

    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        names = ", ".join(path.name for path in matches)
        raise ValueError(f"Multiple workbooks contain the required sheets: {names}. Set {env_var} to choose one.")

    raise FileNotFoundError(f"Could not find a workbook containing sheets: {', '.join(required_sheets)} in {ARTIFACT_DIR}")


def load_private_import_config() -> dict:
    """Load customer-specific names and site mappings without committing them."""
    config_path = os.environ.get("JMI_IMPORT_CONFIG_FILE")
    if not config_path:
        return {"driver_names": [], "location_regions": {}}

    path = Path(config_path).expanduser()
    payload = json.loads(path.read_text())
    if not isinstance(payload, dict):
        raise ValueError("JMI_IMPORT_CONFIG_FILE must contain a JSON object")

    driver_names = payload.get("driver_names", [])
    location_regions = payload.get("location_regions", {})
    if not isinstance(driver_names, list) or not isinstance(location_regions, dict):
        raise ValueError("Import config must contain driver_names (array) and location_regions (object)")

    return {
        "driver_names": [canonical(name) for name in driver_names if clean(name)],
        "location_regions": {canonical(name): clean(region) for name, region in location_regions.items() if clean(name) and clean(region)},
    }


def sheet_by_name(wb, name: str):
    expected = name.strip().lower()
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() == expected:
            return wb[sheet_name]
    return None


def required_sheet(wb, name: str):
    sheet = sheet_by_name(wb, name)
    if sheet is None:
        raise KeyError(f"Required sheet '{name}' not found. Available sheets: {', '.join(wb.sheetnames)}")
    return sheet


def optional_sheet(wb, name: str):
    return sheet_by_name(wb, name)


def value_at(values, index):
    return values[index] if index < len(values) else None


def mapped_value(values, columns, key):
    return value_at(values, columns[key])


def region_for(location: str | None) -> str:
    if not location:
        return "Unknown"
    loc = canonical(location)
    region_map = IMPORT_CONFIG.get("location_regions", {})
    if loc in region_map:
        return region_map[loc]
    parts = [p.strip() for p in re.split(r"/|,|-", loc) if p.strip()]
    mapped_regions = {region_map[part] for part in parts if part in region_map}
    if len(mapped_regions) == 1:
        return mapped_regions.pop()
    return "Islandwide"


def normalize_status(status: str | None, wo_number=None) -> str:
    if str(wo_number).strip().upper() == "PM":
        return "pm"
    s = canonical(status)
    if "WAITING" in s and "PART" in s:
        return "waiting_for_parts"
    if "SCHEDULE" in s:
        return "scheduled"
    if "APPROV" in s:
        return "approved"
    if "ASSESS" in s:
        return "needs_assessment"
    if s == "PM":
        return "pm"
    if s == "CM":
        return "approved"
    if "FABRICAT" in s:
        return "waiting_for_parts"
    return "new"


def priority_from_status(status: str | None) -> str:
    s = canonical(status)
    for priority in ("P1", "P2", "P3", "P4"):
        if priority in s:
            return priority
    if "LEVEL 1" in s:
        return "P1"
    return "P4"


def infer_trade(description: str | None) -> str:
    d = (description or "").lower()
    if any(w in d for w in ["faucet", "sink", "toilet", "drain", "p-trap", "water", "leak"]):
        return "Plumbing"
    if any(w in d for w in ["ac", "airconditioning", "freezer", "cooler", "refrigeration", "ice machine", "condense"]):
        return "HVAC"
    if any(w in d for w in ["electrical", "outlet", "wiring", "light", "generator", "smoke detector", "panel", "led"]):
        return "Electrical"
    if any(w in d for w in ["paint", "painting", "rust"]):
        return "Painting"
    if any(w in d for w in ["door", "cabinet", "counter", "tile", "fabricate", "plexiglass", "shutter", "wall"]):
        return "Carpentry"
    if "landscap" in d:
        return "Landscaping"
    if "pest" in d:
        return "General"
    return "General"


def title_from_description(description: str | None) -> str:
    if not description:
        return "Work order"
    return description[:72] + ("..." if len(description) > 72 else "")


def parse_technicians(wb):
    ws = required_sheet(wb, "Team")
    techs = []
    section = "Mobil"
    for row in ws.iter_rows(values_only=True):
        name = clean(row[0] if len(row) > 0 else None)
        trade = clean(row[1] if len(row) > 1 else None)
        if not name:
            continue
        upper = name.upper()
        if upper in {"TECHNICIAN", "MOBIL", "HOTEL/ KITCHEN / RESTAURANT (HKR)"}:
            if "HOTEL" in upper:
                section = "HKR"
            continue
        if not trade:
            continue
        skills = []
        for raw in re.split(r"/|,", trade.upper()):
            raw = raw.strip()
            if raw in SKILL_ALIASES:
                skills.append(SKILL_ALIASES[raw])
        if not skills:
            skills = [trade.title()]
        techs.append({
            "name": name,
            "primary_trade": skills[0],
            "skills": sorted(set(skills)),
            "is_driver": upper in IMPORT_CONFIG.get("driver_names", []),
            "active": True,
            "division": section,
        })
    return techs


def work_order_record(client, location, wo_number, description, status, source, source_reference, scheduled_date=None, team_name=None, notes=None):
    original_status = clean(status) or ""
    priority = priority_from_status(original_status)
    return {
        "client": client,
        "location": clean(location) or "Unknown",
        "region": region_for(location),
        "external_id": str(wo_number).strip() if wo_number else None,
        "source": source,
        "source_reference": source_reference,
        "title": title_from_description(clean(description)),
        "description": clean(description),
        "priority": priority,
        "normalized_priority": priority,
        "status": normalize_status(original_status, wo_number),
        "original_status_text": original_status,
        "trade_category": infer_trade(description),
        "scheduled_date": scheduled_date,
        "team_name": clean(team_name),
        "notes": notes,
    }


def parse_mobil_schedule(wb):
    ws = required_sheet(wb, "May2026")
    work_orders = []
    current_date = None
    headers = None
    for row in ws.iter_rows(values_only=True):
        values = [clean(v) for v in row]
        if isinstance(row[0], datetime):
            current_date = row[0].date().isoformat()
        if values[0] == "PROJECT":
            headers = values
            continue
        if not headers or not values[0] or not values[0].upper().startswith("MOBIL"):
            continue
        rec = {h: values[i] if i < len(values) else None for i, h in enumerate(headers) if h}
        description = rec.get("DESCRIPTION")
        if not description or str(rec.get("WO#") or "").strip().upper() == "PM":
            continue
        work_orders.append(work_order_record(
            client="Mobil",
            location=rec.get("LOCATION"),
            wo_number=rec.get("WO#"),
            description=description,
            status=rec.get("STATUS"),
            source="mobil_schedule_import",
            source_reference="Private schedule workbook / daily schedule",
            scheduled_date=current_date,
            team_name=rec.get("TECH ASSIGNED"),
            notes="Normalized from the private schedule workbook.",
        ))
    return work_orders


def parse_approved_work_orders(wb):
    ws = optional_sheet(wb, "Approved Work Orders")
    if ws is None:
        return []

    work_orders = []
    for row in ws.iter_rows(values_only=True):
        values = [clean(v) for v in row]
        client = mapped_value(values, APPROVED_WORK_ORDER_COLUMNS, "client")
        if not client or not client.upper().startswith("MOBIL"):
            continue
        description = mapped_value(values, APPROVED_WORK_ORDER_COLUMNS, "description")
        if not description:
            continue
        status = (
            mapped_value(values, APPROVED_WORK_ORDER_COLUMNS, "dispatch_status")
            or mapped_value(values, APPROVED_WORK_ORDER_COLUMNS, "approval_status")
            or "APPROVED"
        )
        work_orders.append(work_order_record(
            client=client.title(),
            location=mapped_value(values, APPROVED_WORK_ORDER_COLUMNS, "location"),
            wo_number=mapped_value(values, APPROVED_WORK_ORDER_COLUMNS, "wo_number"),
            description=description,
            status=status,
            source="approved_work_orders_import",
            source_reference="Private schedule workbook / approved work orders",
            team_name=mapped_value(values, APPROVED_WORK_ORDER_COLUMNS, "team_name"),
            notes="Normalized approved/material-prep work from the private workbook.",
        ))
    return work_orders


def parse_pm_schedule(wb, source_file):
    ws = required_sheet(wb, "Sheet1")
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 6:
        raise ValueError(f"Sheet1 in {source_file} has only {len(rows)} row(s); expected station headers at row 4 and PM data from row 6.")
    location_row = [clean(v) for v in rows[3]]
    if len(location_row) < 3 or not any(location_row[2:]):
        raise ValueError(f"Sheet1 in {source_file} is missing station headers on row 4.")
    tasks = []
    for row in rows[5:]:
        item = clean(row[0])
        task = clean(row[1])
        if not item or not task:
            continue
        for idx, cell in enumerate(row[2:], start=2):
            if isinstance(cell, datetime) and idx < len(location_row):
                location = location_row[idx]
                if not location:
                    continue
                tasks.append({
                    "client": "Mobil",
                    "location": location,
                    "region": region_for(location),
                    "task_name": task,
                    "trade_category": infer_trade(task),
                    "frequency": "monthly",
                    "scheduled_date": cell.date().isoformat(),
                    "source_file": source_file,
                })
    return tasks


def parse_mobil_embedded_pms(wb):
    ws = optional_sheet(wb, "May2026")
    if ws is None:
        return []

    pms = []
    current_date = None
    headers = None
    for row in ws.iter_rows(values_only=True):
        values = [clean(v) for v in row]
        if isinstance(row[0], datetime):
            current_date = row[0].date().isoformat()
        if values[0] == "PROJECT":
            headers = values
            continue
        if not headers or not values[0] or not values[0].upper().startswith("MOBIL"):
            continue
        rec = {h: values[i] if i < len(values) else None for i, h in enumerate(headers) if h}
        if str(rec.get("WO#") or "").strip().upper() != "PM":
            continue
        pms.append({
            "client": "Mobil",
            "location": rec.get("LOCATION") or "Unknown",
            "region": region_for(rec.get("LOCATION")),
            "task_name": rec.get("DESCRIPTION") or "Preventive maintenance",
            "trade_category": infer_trade(rec.get("DESCRIPTION")),
            "frequency": "monthly",
            "scheduled_date": current_date,
            "source_file": "Private schedule workbook / daily schedule",
        })
    return pms


def parse_typhoon_routes(wb):
    ws = optional_sheet(wb, "Typhoon")
    if ws is None:
        return []
    routes = []
    for col in range(2, ws.max_column + 1, 2):
        team = clean(ws.cell(row=2, column=col).value)
        if not team:
            continue
        locations = []
        for row in range(3, ws.max_row + 1):
            location = clean(ws.cell(row=row, column=col).value)
            if location:
                locations.append(location)
        if locations:
            routes.append({"team": team, "locations": locations})
    return routes


def team_names_from_orders(orders):
    counts = Counter(o.get("team_name") for o in orders if o.get("team_name"))
    teams = []
    for name, _count in counts.most_common(16):
        members = [m.strip() for m in re.split(r"/", name) if m.strip()]
        teams.append({"name": name, "members": members, "region_preference": None})
    return teams


def add_missing_team_technicians(technicians, teams):
    existing = {tech["name"].upper() for tech in technicians}
    missing = sorted({member.upper() for team in teams for member in team["members"] if member.upper() not in existing})
    for name in missing:
        technicians.append({
            "name": name,
            "primary_trade": "General",
            "skills": ["General"],
            "is_driver": name in IMPORT_CONFIG.get("driver_names", []),
            "active": True,
            "division": "Mobil",
        })


def public_safe_data(data):
    """Replace customer-operational details while preserving scheduler shape."""
    technician_names = sorted({tech["name"] for tech in data["technicians"]})
    technician_aliases = {name: f"Technician {index:02d}" for index, name in enumerate(technician_names, start=1)}

    location_records = []
    for record in data["work_orders"] + data["pm_tasks"]:
        key = (record.get("client") or "Demo", record.get("location") or "Unknown")
        location_records.append((key, record.get("region") or "Unknown"))
    for route in data.get("typhoon_routes", []):
        for location in route.get("locations", []):
            location_records.append((("Demo", location), region_for(location)))

    location_aliases = {}
    region_counts = Counter()
    for key, region in sorted(set(location_records), key=lambda item: (item[1], item[0][0], item[0][1])):
        region_counts[region] += 1
        location_aliases[key] = f"{region} Demo Site {region_counts[region]:02d}"

    def client_alias(name):
        value = (name or "").lower()
        if "school" in value or "sodexo" in value:
            return "Schools Demo"
        if "hotel" in value or "kitchen" in value or "restaurant" in value or "hkr" in value:
            return "HKR Demo"
        return "Mobil Demo"

    teams = []
    team_aliases = {}
    for index, team in enumerate(data["teams"], start=1):
        alias = f"Crew {index:02d}"
        team_aliases[team["name"]] = alias
        teams.append({
            **team,
            "name": alias,
            "members": [technician_aliases.get(member, member) for member in team.get("members", [])],
        })

    technicians = []
    for technician in data["technicians"]:
        technicians.append({
            **technician,
            "name": technician_aliases[technician["name"]],
            "division": "Primary" if technician.get("division") == "Mobil" else "Secondary",
        })

    work_orders = []
    for index, work_order in enumerate(data["work_orders"], start=1):
        trade = work_order.get("trade_category") or "General"
        raw_location_key = (work_order.get("client") or "Demo", work_order.get("location") or "Unknown")
        description = PUBLIC_DESCRIPTIONS.get(trade, PUBLIC_DESCRIPTIONS["General"])
        work_orders.append({
            **work_order,
            "client": client_alias(work_order.get("client")),
            "location": location_aliases[raw_location_key],
            "external_id": f"DEMO-WO-{index:04d}",
            "source": "demo_import",
            "source_reference": "Local private source artifact (not committed)",
            "title": f"Demo {trade} work order {index:03d}",
            "description": description,
            "original_status_text": (work_order.get("status") or "new").replace("_", " ").title(),
            "team_name": team_aliases.get(work_order.get("team_name")),
            "notes": "Public-safe synthetic record preserving source workflow shape.",
        })

    pm_tasks = []
    for task in data["pm_tasks"]:
        raw_location_key = (task.get("client") or "Demo", task.get("location") or "Unknown")
        pm_tasks.append({
            **task,
            "client": client_alias(task.get("client")),
            "location": location_aliases[raw_location_key],
            "task_name": f"Demo {task.get('trade_category') or 'General'} preventive maintenance",
            "source_file": "Local private PM source (not committed)",
        })

    routes = []
    for index, route in enumerate(data.get("typhoon_routes", []), start=1):
        routes.append({
            "team": f"Storm Crew {index:02d}",
            "locations": [
                location_aliases.get(("Demo", location), f"{region_for(location)} Demo Route Site")
                for location in route.get("locations", [])
            ],
        })

    return {
        **data,
        "sanitization_version": PUBLIC_SANITIZATION_VERSION,
        "notes": "Public-safe synthetic demo data preserving counts, priorities, statuses, trades, regions, dates, and crew sizes from local private artifacts.",
        "source_files": [],
        "typhoon_routes": routes,
        "technicians": technicians,
        "teams": teams,
        "work_orders": work_orders,
        "pm_tasks": pm_tasks,
    }


def write_public_seed(data):
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(data, indent=2))
    print(f"Wrote {OUT}")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--resanitize-existing",
        action="store_true",
        help="Reapply the current public-safe transform to the committed seed without private source files.",
    )
    args = parser.parse_args()

    if args.resanitize_existing:
        write_public_seed(public_safe_data(json.loads(OUT.read_text())))
        return

    global IMPORT_CONFIG
    IMPORT_CONFIG = load_private_import_config()
    mobil_path = find_required_workbook(["Team", "May2026"], env_var="JOHN_MOBIL_SCHEDULE_FILE")
    pm_path = find_required_workbook(["Sheet1"], env_var="JOHN_PM_SCHEDULE_FILE", excluded_sheets=["Team"])
    mobil_wb = load_workbook(mobil_path, data_only=True)
    pm_wb = load_workbook(pm_path, data_only=True)

    technicians = parse_technicians(mobil_wb)
    work_orders = parse_mobil_schedule(mobil_wb) + parse_approved_work_orders(mobil_wb)
    pm_tasks = parse_pm_schedule(pm_wb, pm_path.name) + parse_mobil_embedded_pms(mobil_wb)
    teams = team_names_from_orders(work_orders)
    add_missing_team_technicians(technicians, teams)

    source_data = {
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "notes": "Private normalized source data. This object must be sanitized before it is written.",
        "source_files": sorted(p.name for p in ARTIFACT_DIR.iterdir() if p.is_file()),
        "typhoon_routes": parse_typhoon_routes(pm_wb),
        "technicians": technicians,
        "teams": teams,
        "work_orders": work_orders,
        "pm_tasks": pm_tasks,
    }
    data = public_safe_data(source_data)
    write_public_seed(data)
    print(f"technicians={len(technicians)} teams={len(teams)} work_orders={len(work_orders)} pm_tasks={len(pm_tasks)}")


if __name__ == "__main__":
    main()
